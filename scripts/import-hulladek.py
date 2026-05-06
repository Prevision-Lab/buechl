#!/usr/bin/env python3
"""
Hulladék katalógus importálása Excel fájlból Directusba.
Forrás: dokumentumok/Hulladék lista_honlapra (003) másolata.xlsx
"""

import openpyxl
import re
import requests
import json
import sys
import time

DIRECTUS_URL = "https://buchl-admin.previsionlab.hu"
DIRECTUS_TOKEN = "8Y9Xh6bvU3Q0Bu4vo7I_e9lJyjypCD_8"
EXCEL_PATH = "/Users/c001os/Dev/buchl-app/dokumentumok/Hulladék lista_honlapra (003) másolata.xlsx"

HEADERS = {
    "Authorization": f"Bearer {DIRECTUS_TOKEN}",
    "Content-Type": "application/json",
}

# ──────────────────────────────────────────
# Segédfüggvények
# ──────────────────────────────────────────

def normalize_code(code: str) -> str:
    """Normalizálja a hulladékkódot: '070213*' → '07 02 13*'"""
    code = code.strip()
    # Ha van csillag, kivesszük, majd visszaadjuk
    hazardous = code.endswith('*')
    digits_only = code.rstrip('*').replace(' ', '')
    if len(digits_only) == 6:
        formatted = f"{digits_only[0:2]} {digits_only[2:4]} {digits_only[4:6]}"
        return formatted + ('*' if hazardous else '')
    return code

def is_waste_code(val: str) -> bool:
    """Igaz, ha a string hulladékkódnak tűnik."""
    return bool(re.match(r'^\s*\d{2}[\s]?\d{2}[\s]?\d{2}\*?\s*$', val))

def directus_create_item(collection: str, item: dict) -> dict:
    """Létrehoz egy elemet a Directus-ban."""
    resp = requests.post(
        f"{DIRECTUS_URL}/items/{collection}",
        headers=HEADERS,
        json=item,
    )
    if not resp.ok:
        print(f"  HIBA ({resp.status_code}): {resp.text[:200]}", file=sys.stderr)
        resp.raise_for_status()
    return resp.json()["data"]

def directus_create_items_bulk(collection: str, items: list) -> list:
    """Tömegesen hoz létre elemeket (max 100 darab / hívás)."""
    resp = requests.post(
        f"{DIRECTUS_URL}/items/{collection}",
        headers=HEADERS,
        json=items,
    )
    if not resp.ok:
        print(f"  HIBA ({resp.status_code}): {resp.text[:300]}", file=sys.stderr)
        resp.raise_for_status()
    return resp.json()["data"]

# ──────────────────────────────────────────
# Excel elemzés
# ──────────────────────────────────────────

def parse_sheet1(ws):
    """
    Elemzi a 'Telephelyi gyűjtés-előkezelés' lapot.
    Visszaad egy listát szekciókról: {hely, ktj, engedely, items}
    """
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    sections = []
    current = None

    for i, row in enumerate(rows, 1):
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if row[1] else ''

        if not col0:
            continue

        if col0 in ('Hulladék kód', 'Hasznosítás'):
            continue

        if is_waste_code(col0):
            if current:
                code = normalize_code(col0)
                name = col1.replace('\xa0', ' ').strip()
                current['items'].append({'hulladek_kod': code, 'megnevezes': name})
            continue

        # Engedélyszám sor
        if re.match(r'^(GY/|PE/|1221)', col0):
            if current:
                if current['engedely']:
                    current['engedely'] += '; ' + col0.rstrip('.')
                else:
                    current['engedely'] = col0.rstrip('.')
            continue

        # Telephely fejléc sor (pl. "9027 Győr, Csörgőfa sor 8. - KTJ:100882646")
        ktj_match = re.search(r'KTJ[:\s]*(\d+)', col0)
        ktj = ktj_match.group(1) if ktj_match else ''
        hely = re.sub(r'\s*-?\s*KTJ[:\s]*\d+', '', col0).strip().rstrip(' -')

        # Ha azonos helyszín mint az előző szekció → összevonás
        if current and current['hely_raw'] == hely:
            # Maradunk a jelenlegi szekciónál, csak új engedélyszám jöhet
            pass
        else:
            if current:
                sections.append(current)
            current = {
                'hely_raw': hely,
                'ktj': ktj,
                'engedely': '',
                'items': [],
            }

    if current:
        sections.append(current)

    return sections


def parse_sheet2(ws):
    """Elemzi az 'Országos száll-gyűjt-ker' lapot."""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    engedely = str(rows[0][0]).strip() if rows[0][0] else ''
    items = []
    for row in rows[2:]:
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if row[1] else ''
        if col0 and is_waste_code(col0):
            code = normalize_code(col0)
            name = col1.replace('\xa0', ' ').strip()
            items.append({'hulladek_kod': code, 'megnevezes': name})
    return engedely, items


# ──────────────────────────────────────────
# Fő logika
# ──────────────────────────────────────────

def main():
    print("📂 Excel betöltése...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Sheet 1 feldolgozása
    print("\n📋 Sheet 1: Telephelyi gyűjtés-előkezelés")
    sections1 = parse_sheet1(wb['Telephelyi gyűjtés-előkezelés'])

    # Sheet 2 feldolgozása
    print("📋 Sheet 2: Országos száll-gyűjt-ker")
    szallitas_engedely, szallitas_items = parse_sheet2(wb['Országos száll-gyűjt-ker'])

    # Telephelyek mappálása (hely_raw → kulcs/megnevezés)
    TELEPHELY_MAP = {
        'Győr, Csörgőfa sor 8.': {
            'kulcs': 'csorgfa',
            'nev': 'Győr, Csörgőfa sor 8.',
            'nev_rovid': 'Csörgőfa',
        },
        'Győr, Reptéri út 6.': {
            'kulcs': 'repter',
            'nev': 'Győr, Reptéri út 6.',
            'nev_rovid': 'Reptér',
        },
        'Győr, Reptéri út 6.  -válogatómű': {
            'kulcs': 'repter',
            'nev': 'Győr, Reptéri út 6.',
            'nev_rovid': 'Reptér',
        },
        'Győr, Kandó K. u 17.': {
            'kulcs': 'ati',
            'nev': 'Győr, Kandó K. u 17.',
            'nev_rovid': 'ATI',
        },
        'Nyíregyháza, Derkovits u. 132.': {
            'kulcs': 'nyiregyhaza',
            'nev': 'Nyíregyháza, Derkovits u. 132.',
            'nev_rovid': 'Nyíregyháza',
        },
        'Győr, Kandó K. u 10.': {
            'kulcs': 'karitasz',
            'nev': 'Győr, Kandó K. u 10.',
            'nev_rovid': 'Kandó 10.',
        },
    }

    # Telephely adatok összegyűjtése (összevonás azonos kulcs szerint)
    telephely_data = {}

    for sec in sections1:
        hely_raw = sec['hely_raw']
        # Normalizálás
        hely_clean = re.sub(r'\s+', ' ', hely_raw).strip()
        # Strip address prefix (postal code)
        hely_clean = re.sub(r'^\d{4}\s+', '', hely_clean)

        # Keressük a mappában (részleges egyezés is)
        matched_key = None
        for map_key in TELEPHELY_MAP:
            if map_key.lower() in hely_clean.lower() or hely_clean.lower() in map_key.lower():
                matched_key = map_key
                break

        if not matched_key:
            # Próbáljunk KTJ alapján
            print(f"  ⚠️  Ismeretlen telephely: '{hely_clean}' (KTJ: {sec['ktj']})")
            continue

        kulcs = TELEPHELY_MAP[matched_key]['kulcs']

        if kulcs not in telephely_data:
            telephely_data[kulcs] = {
                **TELEPHELY_MAP[matched_key],
                'ktj': sec['ktj'],
                'engedely': sec['engedely'],
                'items': [],
            }
        else:
            # Kombináljuk az engedélyeket
            existing = telephely_data[kulcs]['engedely']
            new_engedely = sec['engedely']
            if new_engedely and new_engedely not in existing:
                telephely_data[kulcs]['engedely'] = existing + ('; ' if existing else '') + new_engedely

        # Elemek hozzáadása (duplikátum szűréssel kód alapján)
        existing_codes = {it['hulladek_kod'] for it in telephely_data[kulcs]['items']}
        for item in sec['items']:
            if item['hulladek_kod'] not in existing_codes:
                telephely_data[kulcs]['items'].append(item)
                existing_codes.add(item['hulladek_kod'])
            else:
                pass  # Duplikátum kihagyva

    # Szállítás telephely
    telephely_data['szallitas'] = {
        'kulcs': 'szallitas',
        'nev': 'Szállítható hulladékok',
        'nev_rovid': 'Szállítás',
        'ktj': '-',
        'engedely': szallitas_engedely,
        'items': szallitas_items,
    }

    # Összefoglalás
    print("\n📊 Összefoglalás:")
    total_items = 0
    for kulcs, data in telephely_data.items():
        cnt = len(data['items'])
        total_items += cnt
        print(f"  {kulcs:15} | {data['nev']:40} | {cnt:4} elem | {data['engedely'][:60]}")
    print(f"\n  Összes elem: {total_items}")

    # ──────────────────────────────────────────
    # Directus feltöltés
    # ──────────────────────────────────────────
    print("\n🚀 Directus feltöltés kezdése...")

    # Rendezés: szállítás utoljára (hogy a telephelyek sorrendje logikus legyen)
    SORREND = ['csorgfa', 'repter', 'ati', 'nyiregyhaza', 'karitasz', 'szallitas']

    telephely_id_map = {}  # kulcs → directus id

    for i, kulcs in enumerate(SORREND, 1):
        if kulcs not in telephely_data:
            print(f"  ⚠️  Hiányzó telephely: {kulcs}")
            continue

        data = telephely_data[kulcs]
        print(f"\n  [{i}/6] Telephely létrehozása: {data['nev']} ({len(data['items'])} elem)")

        telephely_item = {
            'nev': data['nev'],
            'nev_rovid': data['nev_rovid'],
            'kulcs': kulcs,
            'ktj': data['ktj'],
            'engedely': data['engedely'],
            'sort': i,
            'status': 'published',
        }

        created = directus_create_item('hulladek_telephelyek', telephely_item)
        telephely_id = created['id']
        telephely_id_map[kulcs] = telephely_id
        print(f"     ✅ Telephely ID: {telephely_id}")

        # Hulladékelemek feltöltése 100-as kötegekben
        items = data['items']
        batch_size = 100
        total_batches = (len(items) + batch_size - 1) // batch_size

        for batch_num in range(total_batches):
            batch = items[batch_num * batch_size:(batch_num + 1) * batch_size]
            payload = [
                {
                    'hulladek_kod': it['hulladek_kod'],
                    'megnevezes': it['megnevezes'],
                    'telephely': telephely_id,
                    'sort': batch_num * batch_size + j + 1,
                    'status': 'published',
                }
                for j, it in enumerate(batch)
            ]
            directus_create_items_bulk('hulladek_katalogus', payload)
            print(f"     Köteg {batch_num + 1}/{total_batches} feltöltve ({len(batch)} elem)", end='\r')
            time.sleep(0.2)  # Rate limit elkerülés

        print(f"     ✅ {len(items)} elem feltöltve                    ")

    print(f"\n✅ Kész! {total_items} hulladékelem feltöltve {len(telephely_id_map)} telephelyre.")

if __name__ == '__main__':
    main()
